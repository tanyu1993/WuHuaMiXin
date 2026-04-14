
import os, sys
# 1. 模块自适应注入 (Local & Root Glue)
_FILE_DIR = os.path.dirname(os.path.realpath(__file__))
# 递归向上寻找直到发现 part_ 目录作为模块根
_MOD_ROOT = _FILE_DIR
while _MOD_ROOT != os.path.dirname(_MOD_ROOT) and not os.path.basename(_MOD_ROOT).startswith('part_'):
    _MOD_ROOT = os.path.dirname(_MOD_ROOT)

_PROJECT_ROOT = os.path.dirname(_MOD_ROOT)

if _MOD_ROOT not in sys.path: sys.path.insert(0, _MOD_ROOT)
if _PROJECT_ROOT not in sys.path: sys.path.insert(0, _PROJECT_ROOT)

import os, sys

import random
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

class WHMXAggressiveSim:
    def __init__(self):
        # --- 初始状态 ---
        self.inventory_pulls = 60.0
        self.inventory_red_cards = 0.0
        self.inventory_points = 0.0
        self.pity_counter = 0
        self.total_spent_money = 0.0
        self.history = []
        
        self.pending_upgrades = []
        self.peak_pulls = 60.0
        self.max_drawdown = 0.0
        
        # --- B 档精简方案 (~152元/月) ---
        self.INCOME_PULLS_PER_POOL = (127.6 / 30) * 21  # 89.32 抽
        self.INCOME_REDCARDS_PER_POOL = (6.93 / 30) * 21 # 4.85 张
        self.COST_PER_POOL = (151.5 / 30) * 21 # 106 元

    def do_single_pull(self):
        self.pity_counter += 1
        prob = 0.025 if self.pity_counter <= 50 else 0.025 + (self.pity_counter - 50) * 0.05
        if random.random() < prob:
            curr_pity = self.pity_counter
            self.pity_counter = 0
            return True, (random.random() < 0.5), curr_pity
        return False, False, 0

    def run_simulation(self, years=1000):
        pool_global_idx = 0
        for year in range(1, years + 1):
            pools_config = ["general"] * 17
            limited_slots = [2, 5, 14, random.randint(7, 16)]
            for s in limited_slots: pools_config[s-1] = "limited"
            strong_slots = []
            while len(strong_slots) < 5:
                s = random.randint(0, 16)
                if pools_config[s] == "general": pools_config[s] = "strong"; strong_slots.append(s)

            for i, p_type in enumerate(pools_config):
                pool_global_idx += 1
                self.inventory_pulls += self.INCOME_PULLS_PER_POOL
                self.inventory_red_cards += self.INCOME_REDCARDS_PER_POOL
                self.total_spent_money += self.COST_PER_POOL
                
                # 更新最高水位
                if self.inventory_pulls > self.peak_pulls:
                    self.peak_pulls = self.inventory_pulls
                
                # 计算回撤
                drawdown = self.peak_pulls - self.inventory_pulls
                if drawdown > self.max_drawdown:
                    self.max_drawdown = drawdown

                pulls_spent = 0
                up_copies = 0
                total_reds_got = 0
                spooks = 0
                pity_records = []
                
                # --- 抽卡阶段 ---
                if p_type == "limited":
                    while pulls_spent < 160:
                        if self.inventory_pulls < 10: break
                        self.inventory_pulls -= 10
                        pulls_spent += 10
                        for _ in range(10):
                            res, is_up, p_val = self.do_single_pull()
                            if res:
                                total_reds_got += 1
                                if is_up: up_copies += 1; pity_records.append(f"UP({p_val})")
                                else: spooks += 1; pity_records.append(f"歪({p_val})")
                    if pulls_spent >= 160: up_copies += 1; pity_records.append("井(160)")
                else:
                    while True:
                        if self.inventory_pulls < 10: break
                        self.inventory_pulls -= 10
                        pulls_spent += 10
                        found_up = False
                        for _ in range(10):
                            res, is_up, p_val = self.do_single_pull()
                            if res:
                                total_reds_got += 1
                                if is_up: up_copies += 1; found_up = True; pity_records.append(f"UP({p_val})")
                                else: spooks += 1; pity_records.append(f"歪({p_val})")
                        if up_copies >= 1: self.inventory_points += (pulls_spent * 10); break
                        if pulls_spent >= 160: up_copies += 1; self.inventory_points += (pulls_spent * 10); break

                # 记录补强目标
                target = 7 if p_type == "limited" else (4 if p_type == "strong" else 0)
                if target > 0 and up_copies > 0:
                    self.pending_upgrades.append({"pool_idx": pool_global_idx, "type": p_type, "copies": up_copies, "target": target, "supp": 0})

                # --- 月度补强检查 ---
                supp_this_month = 0
                self.pending_upgrades.sort(key=lambda x: (0 if x["type"]=="limited" else 1, x["pool_idx"]))
                for item in self.pending_upgrades:
                    while item["copies"] < item["target"]:
                        if self.inventory_red_cards >= 4:
                            self.inventory_red_cards -= 4; item["copies"] += 1; item["supp"] += 1; supp_this_month += 1
                        elif self.inventory_points >= 4000:
                            self.inventory_points -= 4000; item["copies"] += 1; item["supp"] += 1; supp_this_month += 1
                        else: break
                self.pending_upgrades = [x for x in self.pending_upgrades if x["copies"] < x["target"]]

                self.history.append({
                    "序号": pool_global_idx,
                    "年份": year,
                    "卡池": p_type.replace("limited","限定").replace("strong","强力").replace("general","图鉴"),
                    "期初抽数": round(self.inventory_pulls + pulls_spent, 1),
                    "消耗抽数": pulls_spent,
                    "期末抽数": round(self.inventory_pulls, 1),
                    "当期UP数": up_copies,
                    "本月补强数": supp_this_month,
                    "歪了没": "是" if spooks > 0 else "否",
                    "流水记录": " | ".join(pity_records),
                    "期末红卡": round(self.inventory_red_cards, 1),
                    "期末积分": int(self.inventory_points),
                    "最大回撤": round(self.max_drawdown, 1)
                })

    def save(self, filename):
        df = pd.DataFrame(self.history)
        df.to_excel(filename, index=False)
        from openpyxl import load_workbook
        wb = load_workbook(filename)
        ws = wb.active
        red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        for cell in ws[1]:
            cell.fill = header_fill; cell.font = Font(bold=True); cell.alignment = Alignment(horizontal='center')
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            if row[8].value == "是": row[8].fill = red_fill
        wb.save(filename)
        return self.max_drawdown, self.inventory_pulls

if __name__ == "__main__":
    sim = WHMXAggressiveSim()
    sim.run_simulation(1000)
    max_dd, final_pulls = sim.save("DATA_ASSETS/物华弥新_1000年人生模拟报告_激进ZZ6版.xlsx")
    print(f"模拟完成！1000年最大回撤: {max_dd} 抽")
    print(f"1000年后结余抽数: {final_pulls} 抽")

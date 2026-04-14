"""
物华弥新技能库 Excel 导出工具
功能：将 skill_db.json 转换为格式化的 Excel 表格
特点：
  - 一行一个技能
  - 纯文本，无 Markdown/emoji
  - 保留完整技能描述
  - 标注所属器者
"""

import json
import re
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# === 配置 ===
INPUT_FILE = r'C:\Users\Administrator\projects\WuHuaMiXin\DATA_ASSETS\skill_db.json'
OUTPUT_FILE = r'C:\Users\Administrator\projects\WuHuaMiXin\DATA_ASSETS\物华弥新_技能库.xlsx'

# === 清理函数：去除 Markdown 和 emoji ===
def clean_text(text):
    """清洗技能文本：去除 Markdown，保留中文和必要格式"""
    if not text:
        return ''
    
    # 1. 去除 Markdown 标题标记 (#####)
    text = re.sub(r'#{1,6}\s*', '', text)
    
    # 2. 去除 emoji（仅移除常见图形符号，不动中文）
    # 使用否定前瞻：只移除非中文、非字母、非数字、非标点的字符
    # 更安全的做法：显式保留 CJK、字母、数字、标点、空格、换行
    cleaned = []
    for char in text:
        code = ord(char)
        # 保留：中文 (CJK), 英文, 数字, 常见标点, 空格, 换行
        if (0x4E00 <= code <= 0x9FFF or  # CJK Unified Ideographs
            0x3000 <= code <= 0x303F or  # CJK标点
            0xFF00 <= code <= 0xFFEF or  # 全角字符
            0x000A <= code <= 0x000A or  # \n 换行符
            0x0020 <= code <= 0x007E or  # ASCII (空格到~)
            0x00A0 <= code <= 0x00FF or  # Latin-1补充
            code == 0x200B or             # 零宽空格（特殊处理）
            char in '、。《》【】→·%+-/（）()[]'):
            cleaned.append(char)
        # 其他字符（emoji等）跳过
    
    text = ''.join(cleaned)
    
    # 3. 清理多余空格（但保留换行）
    text = re.sub(r' {2,}', ' ', text)  # 多空格变单空格
    # 不合并空行，保留原始换行结构
    text = text.strip()
    
    return text

# === 分类映射 ===
CATEGORY_MAP = {
    'zhizhi': '致知',
    'active': '主动技能',
    'passive': '被动技能',
    'huanzhang': '焕章'
}

# === 主逻辑 ===
def main():
    print(">>> 正在读取技能库...")
    with open(INPUT_FILE, 'r', encoding='utf-8') as f:
        skill_db = json.load(f)
    
    print(f">>> 器者数量: {len(skill_db)}")
    
    # 展开所有技能为扁平列表
    rows = []
    for char_name in sorted(skill_db.keys()):
        char_skills = skill_db[char_name]
        
        for category, skills in char_skills.items():
            category_name = CATEGORY_MAP.get(category, category)
            
            for skill in skills:
                skill_name = skill.get('name', '')
                description = skill.get('description', '')
                status_links = ', '.join(skill.get('status_links', []))
                
                # 清理文本
                clean_desc = clean_text(description)
                clean_name = clean_text(skill_name)
                
                # 去除描述中开头的技能名（避免重复）
                if clean_desc.startswith(clean_name):
                    clean_desc = clean_desc[len(clean_name):].lstrip('\n ')
                
                # 提取技能类型（从名称中判断）
                skill_type = ''
                if category == 'zhizhi':
                    # 致知：类型统一为"致知"，技能名为"致知 壹/贰/叁..."
                    skill_type = '致知'
                elif '常击' in clean_name:
                    skill_type = '常击'
                elif '职业技能' in clean_name:
                    skill_type = '职业技能'
                elif '绝技' in clean_name:
                    skill_type = '绝技'
                elif '被动' in clean_name:
                    # 被动1、被动2、被动3
                    skill_type = clean_name.split('：')[0].strip() if '：' in clean_name else '被动'
                elif category == 'huanzhang':
                    skill_type = '焕章'
                
                rows.append({
                    '器者名': char_name,
                    '技能分类': category_name,
                    '技能类型': skill_type,
                    '技能名': clean_name,
                    '技能描述': clean_desc,
                    '状态关联': status_links
                })
    
    print(f">>> 技能总数: {len(rows)}")
    
    # === 创建 Excel ===
    print(">>> 正在生成 Excel...")
    wb = Workbook()
    ws = wb.active
    ws.title = "技能库"
    
    # 表头
    headers = ['器者名', '技能分类', '技能类型', '技能名', '技能描述', '状态关联']
    ws.append(headers)
    
    # 数据行
    for row in rows:
        ws.append([
            row['器者名'],
            row['技能分类'],
            row['技能类型'],
            row['技能名'],
            row['技能描述'],
            row['状态关联']
        ])
    
    # === 格式化 ===
    
    # 表头样式
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # 数据行样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 列宽设置
    column_widths = {
        'A': 12,   # 器者名
        'B': 10,   # 技能分类
        'C': 12,   # 技能类型
        'D': 25,   # 技能名
        'E': 80,   # 技能描述（最宽）
        'F': 20    # 状态关联
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # 应用样式到所有数据行
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        # 交替行颜色
        if row_idx % 2 == 0:
            fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        else:
            fill = None
        
        for cell in row:
            if fill:
                cell.fill = fill
            cell.border = thin_border
            
            # 对齐方式
            if cell.column == 5:  # E列：技能描述
                cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
            elif cell.column in [1, 2, 3, 4]:  # A-D列
                cell.alignment = Alignment(wrap_text=False, vertical='center', horizontal='left')
            else:  # F列
                cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    
    # 冻结首行
    ws.freeze_panes = 'A2'
    
    # 自动调整行高（技能描述列）
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        desc_cell = row[4]  # E列
        if desc_cell.value:
            line_count = str(desc_cell.value).count('\n') + 1
            # 基础行高 15，每多一行增加 15
            row_height = max(15, line_count * 15)
            ws.row_dimensions[desc_cell.row].height = row_height
    
    # 保存
    print(f">>> 正在保存: {OUTPUT_FILE}")
    wb.save(OUTPUT_FILE)
    
    print(f">>> ✅ 完成！")
    print(f">>> 文件位置: {OUTPUT_FILE}")
    print(f">>> 总行数: {ws.max_row} (含表头)")

if __name__ == '__main__':
    main()

import random
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

class WHMXFinalSimV6:
    def __init__(self):
        # --- 初始状态 ---
        self.inventory_pulls = 60.0
        self.inventory_red_cards = 0.0
        self.inventory_points = 0.0
        self.pity_counter = 0
        self.total_spent_money = 0.0
        self.history = []
        
        # 待补强清单: [{"pool_idx": idx, "type": type, "current_copies": n}]
        self.pending_upgrades = []
        
        # --- 200元/月 大月卡方案 ---
        self.INCOME_PULLS_PER_POOL = (149.0 / 30) * 21  # 104.3 抽
        self.INCOME_REDCARDS_PER_POOL = (6.93 / 30) * 21 # 4.85 张
        self.COST_PER_POOL = (200.0 / 30) * 21 # 140 元

    def do_single_pull(self):
        self.pity_counter += 1
        prob = 0.025 if self.pity_counter <= 50 else 0.025 + (self.pity_counter - 50) * 0.05
        if random.random() < prob:
            curr_pity = self.pity_counter
            self.pity_counter = 0
            return True, (random.random() < 0.5), curr_pity
        return False, False, 0

    def run_simulation(self, years=1000):
        pool_global_idx = 0
        for year in range(1, years + 1):
            pools_config = ["general"] * 17
            limited_slots = [2, 5, 14, random.choice([8, 11, 16])]
            for s in limited_slots: pools_config[s-1] = "limited"
            strong_slots = []
            while len(strong_slots) < 5:
                s = random.randint(0, 16)
                if pools_config[s] == "general":
                    pools_config[s] = "strong"
                    strong_slots.append(s)

            for i, p_type in enumerate(pools_config):
                pool_global_idx += 1
                ver = (i // 2) + 1
                sub_ver = (i % 2) + 1
                
                init_pulls = self.inventory_pulls
                init_points = self.inventory_points
                init_reds = self.inventory_red_cards
                
                # 周期收入
                self.inventory_pulls += self.INCOME_PULLS_PER_POOL
                self.inventory_red_cards += self.INCOME_REDCARDS_PER_POOL
                self.total_spent_money += self.COST_PER_POOL
                
                pulls_spent = 0
                up_copies = 0
                total_reds_got = 0
                spooks = 0
                pity_records = []
                
                # --- 抽卡阶段 ---
                if p_type == "limited":
                    # 限定池：死磕 160 抽
                    while pulls_spent < 160:
                        if self.inventory_pulls < 10: break # 虽然原则上死磕，但没钱也得停
                        self.inventory_pulls -= 10
                        pulls_spent += 10
                        for _ in range(10):
                            res, is_up, p_val = self.do_single_pull()
                            if res:
                                total_reds_got += 1
                                if is_up: 
                                    up_copies += 1
                                    pity_records.append(f"UP({p_val})")
                                else: 
                                    spooks += 1
                                    pity_records.append(f"歪({p_val})")
                    # 160 领井
                    if pulls_spent >= 160:
                        up_copies += 1
                        pity_records.append("井(160)")
                else:
                    # 强力和图鉴：抽到 1 个即停
                    while True:
                        if self.inventory_pulls < 10: break
                        self.inventory_pulls -= 10
                        pulls_spent += 10
                        found_in_10 = False
                        for _ in range(10):
                            res, is_up, p_val = self.do_single_pull()
                            if res:
                                total_reds_got += 1
                                if is_up: 
                                    up_copies += 1
                                    pity_records.append(f"UP({p_val})")
                                    found_in_10 = True
                                else: 
                                    spooks += 1
                                    pity_records.append(f"歪({p_val})")
                        
                        if up_copies >= 1:
                            # 抽到后，不满 160 给积分，超过 160 也不领井给积分 (按用户策略)
                            self.inventory_points += (pulls_spent * 10)
                            break
                        if pulls_spent >= 160:
                            # 160 必出一个，但不额外吃井，给积分
                            up_copies += 1
                            self.inventory_points += (pulls_spent * 10)
                            break

                # 记录核心器者到待补强清单
                if p_type in ["limited", "strong"] and up_copies > 0:
                    self.pending_upgrades.append({
                        "pool_idx": pool_global_idx,
                        "type": p_type,
                        "copies": up_copies,
                        "supplemented_this_month": 0
                    })

                # --- 月度补强检查阶段 ---
                reds_used_this_month = 0
                points_used_this_month = 0
                
                # 优先级：限定 > 强力，按发布顺序
                self.pending_upgrades.sort(key=lambda x: (0 if x["type"]=="limited" else 1, x["pool_idx"]))
                
                for item in self.pending_upgrades:
                    while item["copies"] < 4:
                        # 4张红卡补一个
                        # 1. 优先用库存红卡
                        if self.inventory_red_cards >= 4:
                            self.inventory_red_cards -= 4
                            item["copies"] += 1
                            item["supplemented_this_month"] += 1
                            reds_used_this_month += 4
                        # 2. 红卡不足，用积分换 (4000分 = 1本体)
                        elif self.inventory_points >= 4000:
                            self.inventory_points -= 4000
                            item["copies"] += 1
                            item["supplemented_this_month"] += 1
                            points_used_this_month += 4000
                        else:
                            break
                
                # 清理已达标的角色
                self.pending_upgrades = [x for x in self.pending_upgrades if x["copies"] < 4]

                # 记录历史
                self.history.append({
                    "序号": pool_global_idx,
                    "年份": year,
                    "版本": f"v{ver}.{sub_ver}",
                    "卡池类型": p_type.replace("limited","限定").replace("strong","强力").replace("general","图鉴"),
                    "初始抽数": round(init_pulls, 1),
                    "收入抽数": round(self.INCOME_PULLS_PER_POOL, 1),
                    "消耗抽数": pulls_spent,
                    "期末抽数": round(self.inventory_pulls, 1),
                    "当前UP数": up_copies,
                    "本月红卡补位": sum([x["supplemented_this_month"] for x in self.history if x.get("序号") == pool_global_idx]), # 占位
                    "最终致知": 0, # 稍后处理
                    "消耗红卡": reds_used_this_month,
                    "消耗积分": points_used_this_month,
                    "歪了没": "是" if spooks > 0 else "否",
                    "出货记录": " | ".join(pity_records),
                    "期末红卡": round(self.inventory_red_cards, 1),
                    "期末积分": int(self.inventory_points),
                    "本池氪金": round(self.COST_PER_POOL, 1),
                    "累计氪金": round(self.total_spent_money, 0)
                })
                # 修正：将本月补位数据回填到当月历史
                self.history[-1]["本月红卡补位"] = sum(x["supplemented_this_month"] for x in self.pending_upgrades if x["pool_idx"] == pool_global_idx) # 这是一个近似值

    def save_to_excel(self, filename):
        df = pd.DataFrame(self.history)
        # 补齐“最终致知”列：这在连续模拟中比较难实时追踪，我们直接看是否达标
        df.to_excel(filename, index=False)
        from openpyxl import load_workbook
        wb = load_workbook(filename)
        ws = wb.active
        header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        wb.save(filename)

if __name__ == "__main__":
    sim = WHMXFinalSimV6()
    sim.run_simulation(1000)
    sim.save_to_excel("whmx/物华弥新_1000年人生模拟报告_终极策略版.xlsx")
    print("报告已生成！")

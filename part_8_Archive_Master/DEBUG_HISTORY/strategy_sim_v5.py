import random
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

class WHMXFinalSimV5:
    def __init__(self):
        # --- 初始状态 ---
        self.inventory_pulls = 60.0
        self.inventory_red_cards = 0.0
        self.inventory_points = 0.0
        self.pity_counter = 0
        self.total_spent_money = 0.0
        self.history = []
        
        # --- 200元/月 大月卡方案 ---
        self.INCOME_PULLS_PER_POOL = (149.0 / 30) * 21  # 104.3 抽
        self.INCOME_REDCARDS_PER_POOL = (6.93 / 30) * 21 # 4.85 张
        self.COST_PER_POOL = (200.0 / 30) * 21 # 140 元

    def do_single_pull(self):
        self.pity_counter += 1
        prob = 0.025 if self.pity_counter <= 50 else 0.025 + (self.pity_counter - 50) * 0.05
        if random.random() < prob:
            curr_pity = self.pity_counter
            self.pity_counter = 0
            return True, (random.random() < 0.5), curr_pity
        return False, False, 0

    def run_simulation(self, years=1000):
        pool_global_idx = 0
        for year in range(1, years + 1):
            pools_config = ["general"] * 17
            limited_slots = [2, 5, 14, random.choice([8, 11, 16])]
            for s in limited_slots: pools_config[s-1] = "limited"
            strong_slots = []
            while len(strong_slots) < 5:
                s = random.randint(0, 16)
                if pools_config[s] == "general":
                    pools_config[s] = "strong"
                    strong_slots.append(s)

            year_targets_status = [] # 记录当年限定/强力的获取情况

            for i, p_type in enumerate(pools_config):
                pool_global_idx += 1
                ver = (i // 2) + 1
                sub_ver = (i % 2) + 1
                
                init_pulls = self.inventory_pulls
                init_points = self.inventory_points
                init_reds = self.inventory_red_cards
                
                self.inventory_pulls += self.INCOME_PULLS_PER_POOL
                self.inventory_red_cards += self.INCOME_REDCARDS_PER_POOL
                self.total_spent_money += self.COST_PER_POOL
                
                pulls_spent = 0
                up_copies = 0
                total_reds_got = 0
                spooks = 0
                pity_records = []
                
                target = 4 if (p_type == "limited" or p_type == "strong") else 1
                
                while True:
                    if self.inventory_pulls < 10: break
                    self.inventory_pulls -= 10
                    pulls_spent += 10
                    for _ in range(10):
                        res, is_up, p_val = self.do_single_pull()
                        if res:
                            total_reds_got += 1
                            if is_up: 
                                up_copies += 1
                                pity_records.append(f"UP({p_val})")
                            else: 
                                spooks += 1
                                pity_records.append(f"歪({p_val})")
                    
                    # 停止逻辑
                    if p_type == "limited" or p_type == "strong":
                        # 修改：限定也不再死磕，160即停
                        if up_copies >= target:
                            self.inventory_points += (pulls_spent * 10)
                            break
                        if pulls_spent >= 160:
                            up_copies += 1
                            pity_records.append("井(160)")
                            break
                    else: # general
                        if up_copies >= 1:
                            self.inventory_points += (pulls_spent * 10)
                            break
                        if pulls_spent >= 160:
                            up_copies += 1
                            self.inventory_points += (pulls_spent * 10)
                            break

                # 如果是核心器者，记录其缺口
                supplemented = 0
                if p_type in ["limited", "strong"]:
                    year_targets_status.append({"idx": pool_global_idx, "type": p_type, "copies": up_copies})

                self.history.append({
                    "序号": pool_global_idx,
                    "年份": year,
                    "版本": f"v{ver}.{sub_ver}",
                    "卡池类型": p_type.replace("limited","限定").replace("strong","强力").replace("general","图鉴"),
                    "初始抽数": round(init_pulls, 1),
                    "收入抽数": round(self.INCOME_PULLS_PER_POOL, 1),
                    "消耗抽数": pulls_spent,
                    "期末抽数": round(self.inventory_pulls, 1),
                    "出红总数": total_reds_got,
                    "UP个数": up_copies,
                    "红卡补位": 0, # 初始为0，稍后回填
                    "最终致知": up_copies, # 稍后回填
                    "歪了没": "是" if spooks > 0 else "否",
                    "出货记录": " | ".join(pity_records),
                    "达成目标": "√" if up_copies >= target else "×",
                    "当前红卡": round(self.inventory_red_cards, 1),
                    "当前积分": int(self.inventory_points),
                    "本池氪金": round(self.COST_PER_POOL, 1),
                    "累计氪金": round(self.total_spent_money, 0)
                })

            # --- 年末资源补强阶段 ---
            # 1. 优先补限定
            for item in sorted(year_targets_status, key=lambda x: x["type"]): # 字母序 limited 优先于 strong
                gap = 4 - item["copies"]
                if gap > 0:
                    # 将积分转为红卡
                    while self.inventory_points >= 1000:
                        self.inventory_points -= 1000
                        self.inventory_red_cards += 1
                    
                    # 检查红卡是否足够补一个本体 (4张)
                    can_fill = int(self.inventory_red_cards // 4)
                    to_fill = min(gap, can_fill)
                    
                    if to_fill > 0:
                        self.inventory_red_cards -= (to_fill * 4)
                        # 更新历史记录中的该器者数据
                        h_idx = item["idx"] - 1
                        self.history[h_idx]["红卡补位"] = to_fill
                        self.history[h_idx]["最终致知"] = self.history[h_idx]["UP个数"] + to_fill
                        if self.history[h_idx]["最终致知"] >= 4:
                            self.history[h_idx]["达成目标"] = "√"

    def save_to_excel(self, filename):
        df = pd.DataFrame(self.history)
        df.to_excel(filename, index=False)
        from openpyxl import load_workbook
        wb = load_workbook(filename)
        ws = wb.active
        red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            if row[11].value == "是": row[11].fill = red_fill
            if row[14].value == "√": row[14].fill = green_fill
            else: row[14].fill = red_fill
        wb.save(filename)

if __name__ == "__main__":
    sim = WHMXFinalSimV5()
    sim.run_simulation(1000)
    sim.save_to_excel("whmx/物华弥新_1000年人生模拟报告_新策略版.xlsx")
    print("报告已生成！")

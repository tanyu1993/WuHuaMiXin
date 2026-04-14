"""
物华弥新 千年抽卡策略模拟 v12（2026-04-13 更新）
=====================================================
v12 相比 v11 的修复/变更：
  1. 修复：pity_log 记录的保底数是命中后归零的值（全为0的bug）
           → 现在记录命中时的实际抽数（hit_at 变量）
  2. 修复：_pull_limited 漏掉积分累积
           → 限定池每抽同样 +10 积分
  3. 修复：初始库存由 60 改为 200 抽（反映憨总当前账号实际存量）
  4. 其余参数与 v11 完全一致

【版本参数（v11 已修正，v12 沿用）】：
  - 限定版本免费抽：80 抽
  - 其余版本：60 抽
  - 版本结构：8大版本（各2小版本21天）+ 1过渡（28天）= 364天
  - 限定固定位置：大1-上（春节）、大2-上（周年庆）、大5-上（国庆半周年庆）
                 + 第4个随机

【抽卡策略 v12（激进ZZ6版，沿用 v10/v11）】：
  - 限定器者：必抽 160 抽吃井，目标 ZZ6（7本体）
  - 强力器者：抽到 1 个 UP 即停，目标 ZZ3
  - 一般/图鉴：抽到 1 个 UP 即停，目标 ZZ0-1

【B档订阅（152元/月）】：
  大月卡 + 普通月卡 + 增效月卡 + 活动月卡
"""

import os, sys, random
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

# ============================================================
# 参数配置
# ============================================================

# 版本免费抽（v11 修正）
FREE_PULLS_LIMITED  = 80   # 限定器者所在小版本
FREE_PULLS_NORMAL   = 60   # 普通小版本
FREE_PULLS_TRANS    = 60   # 过渡版本

# 版本天数
DAYS_SMALL = 21
DAYS_TRANS = 28

# B档订阅月均（沿用历史校准值，每月30天折算）
SUBS_MONTHLY_PULLS = 41.9   # 抽/月
SUBS_MONTHLY_REDS  = 2.43   # 红卡/月（增效+活动月卡直给）

# 红票红卡（保守=0；如果账号开始溢出，可调整至 0-4）
REDTICKET_PER_MONTH = 0.0   # 每月红票换红卡数量

# 积分换红卡：4000分 = 1张（在代码中从积分池扣除）
POINTS_PER_REDCARD = 4000

# 日均折算
SUBS_DAILY_PULLS = SUBS_MONTHLY_PULLS / 30
SUBS_DAILY_REDS  = SUBS_MONTHLY_REDS  / 30

# 月费
MONTHLY_COST = 152.0


# ============================================================
# 抽卡模型
# ============================================================
def do_single_pull(pity):
    """返回 (命中特出, 是否UP, 新保底计数)"""
    pity += 1
    prob = 0.025 if pity <= 50 else min(1.0, 0.025 + (pity - 50) * 0.05)
    if random.random() < prob:
        return True, (random.random() < 0.5), 0
    return False, False, pity


# ============================================================
# 年度版本结构生成
# ============================================================
def build_year_structure(rng):
    """
    17个周期，每个元素：
      {"days": int, "free_pulls": int, "is_limited": bool,
       "is_strong": bool, "name": str}
    
    限定位置（0-indexed）：
      0  = 大1-上（春节）
      2  = 大2-上（周年庆4月19日附近）
      8  = 大5-上（国庆/半周年庆）
      第4个 = 随机取剩余普通小版本（不含过渡=索引16）
    
    强力器者：5个，随机取剩余普通位置
    其余：图鉴/一般器者
    """
    # 初始化17个普通周期
    structure = []
    for big in range(8):
        structure.append({"days": DAYS_SMALL, "free_pulls": FREE_PULLS_NORMAL,
                          "is_limited": False, "is_strong": False,
                          "name": "大%d-上" % (big + 1)})
        structure.append({"days": DAYS_SMALL, "free_pulls": FREE_PULLS_NORMAL,
                          "is_limited": False, "is_strong": False,
                          "name": "大%d-下" % (big + 1)})
    structure.append({"days": DAYS_TRANS, "free_pulls": FREE_PULLS_TRANS,
                      "is_limited": False, "is_strong": False,
                      "name": "过渡"})

    # 固定限定位置
    fixed_limited = [0, 2, 8]
    available_for_fourth = [i for i in range(16) if i not in fixed_limited]
    fourth_limited = rng.choice(available_for_fourth)
    all_limited = fixed_limited + [fourth_limited]

    for idx in all_limited:
        structure[idx]["is_limited"] = True
        structure[idx]["free_pulls"] = FREE_PULLS_LIMITED

    # 强力器者：从剩余普通版本（非限定，非过渡）随机选5个
    available_for_strong = [i for i in range(16) if i not in all_limited]
    strong_slots = rng.sample(available_for_strong, min(5, len(available_for_strong)))
    for idx in strong_slots:
        structure[idx]["is_strong"] = True

    return structure


# ============================================================
# 主模拟类
# ============================================================
class WHMXSimV11:
    def __init__(self, redticket_per_month=REDTICKET_PER_MONTH):
        # 库存（初始200抽，反映憨总当前账号实际存量）
        self.inventory_pulls  = 200.0
        self.inventory_reds   = 0.0
        self.inventory_points = 0.0
        self.pity             = 0
        self.pity_guaranteed  = False  # 大保底状态（歪了下次必出UP）

        # 统计
        self.total_spent_money = 0.0
        self.peak_pulls        = 200.0
        self.max_drawdown      = 0.0
        self.history           = []

        # 补强队列：{"pool_idx", "type", "copies", "target", "supp"}
        self.pending_upgrades  = []
        self.pool_global_idx   = 0

        # 红票红卡（每月额外上限）
        self.redticket_per_month = redticket_per_month

    def _do_pull(self):
        """执行单抽，返回 (命中特出, 是否UP)"""
        self.pity += 1
        prob = 0.025 if self.pity <= 50 else min(1.0, 0.025 + (self.pity - 50) * 0.05)
        if random.random() < prob:
            self.pity = 0
            if self.pity_guaranteed or random.random() < 0.5:
                self.pity_guaranteed = False
                return True, True
            else:
                self.pity_guaranteed = True
                return True, False
        return False, False

    def _monthly_income(self, days):
        """计算某周期的收入（按天折算）"""
        pulls = SUBS_DAILY_PULLS * days
        reds  = SUBS_DAILY_REDS  * days
        # 红票红卡按周期占月度比例发放（21天≈0.7个月）
        reds += self.redticket_per_month * (days / 30)
        cost  = MONTHLY_COST * (days / 30)
        return pulls, reds, cost

    def _pull_limited(self):
        """限定池：打满160抽，吃井"""
        spent = 0
        up_copies = 0
        reds_got  = 0
        pity_log  = []

        while spent < 160:
            if self.inventory_pulls < 1:
                break
            self.inventory_pulls -= 1
            spent += 1
            self.inventory_points += 10  # 每抽+10积分（修复：v11漏掉）
            hit, is_up = self._do_pull()
            if hit:
                hit_at = spent  # 命中时的实际抽数（修复：不能用pity，命中后已归零）
                reds_got += 1
                if is_up:
                    up_copies += 1
                    pity_log.append("UP(%d)" % hit_at)
                else:
                    pity_log.append("歪(%d)" % hit_at)

        # 吃井：打满160抽（无论有没有额外UP）
        if spent >= 160:
            pity_log.append("井(160)")
            if up_copies == 0:
                up_copies = 1  # 保底给1个UP

        return spent, up_copies, reds_got, pity_log

    def _pull_nolimit(self):
        """普通/强力池：抽到1个UP即停"""
        spent = 0
        up_copies = 0
        reds_got  = 0
        pity_log  = []

        while True:
            if self.inventory_pulls < 1:
                break
            self.inventory_pulls -= 1
            spent += 1
            self.inventory_points += 10
            hit, is_up = self._do_pull()
            if hit:
                hit_at = spent  # 命中时的实际抽数（修复：不能用pity，命中后已归零）
                reds_got += 1
                if is_up:
                    up_copies += 1
                    pity_log.append("UP(%d)" % hit_at)
                else:
                    pity_log.append("歪(%d)" % hit_at)
            if up_copies >= 1:
                break
            if spent >= 160:
                up_copies = 1
                pity_log.append("井(160)")
                break

        return spent, up_copies, reds_got, pity_log

    def _supplement_upgrades(self):
        """用红卡/积分补强，优先限定，按先后顺序"""
        self.pending_upgrades.sort(key=lambda x: (0 if x["type"] == "limited" else 1, x["pool_idx"]))
        total_supp = 0
        for item in self.pending_upgrades:
            while item["copies"] < item["target"]:
                if self.inventory_reds >= 1:
                    self.inventory_reds -= 1
                    item["copies"] += 1
                    item["supp"]   += 1
                    total_supp     += 1
                elif self.inventory_points >= POINTS_PER_REDCARD:
                    self.inventory_points -= POINTS_PER_REDCARD
                    item["copies"] += 1
                    item["supp"]   += 1
                    total_supp     += 1
                else:
                    break
        self.pending_upgrades = [x for x in self.pending_upgrades if x["copies"] < x["target"]]
        return total_supp

    def run_simulation(self, years=1000, seed=None):
        rng = random.Random(seed)
        random.seed(seed)  # 全局随机（抽卡用）

        for year in range(1, years + 1):
            structure = build_year_structure(rng)

            # 每年重置年内峰值（用于统计年内回撤）
            year_peak   = self.inventory_pulls
            year_max_dd = 0.0

            for period in structure:
                self.pool_global_idx += 1
                days     = period["days"]
                is_lim   = period["is_limited"]
                is_str   = period["is_strong"]
                p_type   = "limited" if is_lim else ("strong" if is_str else "general")

                # 收入
                inc_pulls, inc_reds, inc_cost = self._monthly_income(days)
                self.inventory_pulls  += inc_pulls + period["free_pulls"]
                self.inventory_reds   += inc_reds
                self.total_spent_money += inc_cost

                # 更新全局峰值 & 全局回撤
                if self.inventory_pulls > self.peak_pulls:
                    self.peak_pulls = self.inventory_pulls
                drawdown = self.peak_pulls - self.inventory_pulls
                if drawdown > self.max_drawdown:
                    self.max_drawdown = drawdown

                # 更新年内峰值 & 年内回撤
                if self.inventory_pulls > year_peak:
                    year_peak = self.inventory_pulls
                year_dd = year_peak - self.inventory_pulls
                if year_dd > year_max_dd:
                    year_max_dd = year_dd

                # 抽卡
                if is_lim:
                    spent, up_copies, reds_got, pity_log = self._pull_limited()
                else:
                    spent, up_copies, reds_got, pity_log = self._pull_nolimit()

                self.inventory_reds += reds_got

                # 登记补强目标
                target = 7 if is_lim else (4 if is_str else 0)
                if target > 0 and up_copies > 0:
                    self.pending_upgrades.append({
                        "pool_idx": self.pool_global_idx,
                        "type":     p_type,
                        "copies":   up_copies,
                        "target":   target,
                        "supp":     0
                    })

                supp_count = self._supplement_upgrades()

                has_spook = any("歪" in s for s in pity_log)

                self.history.append({
                    "序号":    self.pool_global_idx,
                    "年份":    year,
                    "版本":    period["name"],
                    "类型":    p_type.replace("limited","限定").replace("strong","强力").replace("general","图鉴"),
                    "天数":    days,
                    "版本抽":  period["free_pulls"],
                    "消耗抽":  spent,
                    "UP数":    up_copies,
                    "补强数":  supp_count,
                    "是否歪":  "是" if has_spook else "否",
                    "流水":    " | ".join(pity_log) if pity_log else "-",
                    "期末抽":  round(self.inventory_pulls, 1),
                    "期末红卡": round(self.inventory_reds, 2),
                    "期末积分": int(self.inventory_points),
                    "全局最大回撤": round(self.max_drawdown, 1),
                    "年内最大回撤": round(year_max_dd, 1),
                })

    def save_excel(self, filename):
        df = pd.DataFrame(self.history)
        df.to_excel(filename, index=False)

        wb = load_workbook(filename)
        ws = wb.active

        header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        red_fill    = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        lim_fill    = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

        for cell in ws[1]:
            cell.fill      = header_fill
            cell.font      = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        type_col = None
        spook_col = None
        for i, cell in enumerate(ws[1], start=1):
            if cell.value == "类型":   type_col  = i
            if cell.value == "是否歪": spook_col = i

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            if type_col  and row[type_col  - 1].value == "限定": row[type_col - 1].fill = lim_fill
            if spook_col and row[spook_col - 1].value == "是":   row[spook_col - 1].fill = red_fill

        wb.save(filename)
        print("已保存: %s" % filename)
        return self.max_drawdown, self.inventory_pulls, self.inventory_reds


# ============================================================
# 主程序
# ============================================================
if __name__ == "__main__":
    import time
    print("=" * 70)
    print("  物华弥新 千年抽卡策略模拟 v12")
    print("  参数：限定80抽 | 普通60抽 | B档152元/月 | 初始200抽")
    print("  修复：流水保底数显示 | 限定池积分累积 | 初始库存200抽")
    print("=" * 70)
    print()
    print("【参数校验】")
    free_annual = 4 * FREE_PULLS_LIMITED + 13 * FREE_PULLS_NORMAL
    print("  年度免费抽: 4×%d + 13×%d = %d 抽（月均 %.1f 抽）" % (
        FREE_PULLS_LIMITED, FREE_PULLS_NORMAL, free_annual, free_annual/12))
    print("  订阅月均: %.1f 抽 + %.2f 红卡" % (SUBS_MONTHLY_PULLS, SUBS_MONTHLY_REDS))
    print("  合计月均: %.1f 抽" % (free_annual/12 + SUBS_MONTHLY_PULLS))
    print()

    t0 = time.time()
    sim = WHMXSimV11(redticket_per_month=REDTICKET_PER_MONTH)
    sim.run_simulation(years=1000, seed=42)
    elapsed = time.time() - t0

    print("模拟耗时: %.2f 秒" % elapsed)
    print()
    print("=" * 70)
    print("  千年模拟结果")
    print("=" * 70)
    print()

    h = sim.history
    total_pools  = len(h)
    total_years  = h[-1]["年份"]
    total_money  = sim.total_spent_money

    # 按年统计期末抽（取每年最后一条记录）
    year_end_pulls = {}
    for row in h:
        year_end_pulls[row["年份"]] = row["期末抽"]
    pulls_list = list(year_end_pulls.values())
    pulls_list_sorted = sorted(pulls_list)
    p50  = pulls_list_sorted[int(len(pulls_list_sorted) * 0.50)]
    p05  = pulls_list_sorted[int(len(pulls_list_sorted) * 0.05)]
    p01  = pulls_list_sorted[int(len(pulls_list_sorted) * 0.01)]
    p95  = pulls_list_sorted[int(len(pulls_list_sorted) * 0.95)]

    # 回撤：按年取年内最大回撤（年内独立统计，反映单年最坏运气）
    year_intra_dd = {}
    for row in h:
        y  = row["年份"]
        dd = row["年内最大回撤"]
        if y not in year_intra_dd or dd > year_intra_dd[y]:
            year_intra_dd[y] = dd
    dd_list_sorted = sorted(year_intra_dd.values())
    dd_p50 = dd_list_sorted[int(len(dd_list_sorted) * 0.50)]
    dd_p95 = dd_list_sorted[int(len(dd_list_sorted) * 0.95)]
    dd_p99 = dd_list_sorted[int(len(dd_list_sorted) * 0.99)]
    dd_max = max(dd_list_sorted)

    print("  模拟总年数:   %d 年 (%d 个池子)" % (total_years, total_pools))
    print("  累计月费:     ¥%.0f（月均 ¥%.0f）" % (total_money, MONTHLY_COST))
    print()
    print("  【年末库存抽分布（各年末相对起点的净增）】")
    print("    中位数 (P50): %d 抽" % p50)
    print("    P95（乐观）:  %d 抽" % p95)
    print("    P05（悲观）:  %d 抽" % p05)
    print("    P01（极端）:  %d 抽" % p01)
    print()
    print("  【年内最大回撤分布（每年独立：年内峰值→谷值）】")
    print("    中位数 (P50): %d 抽" % dd_p50)
    print("    五年一遇 (P95): %d 抽（建议安全水位 ≥ 此值）" % dd_p95)
    print("    百年一遇 (P99): %d 抽" % dd_p99)
    print("    历史最大: %d 抽" % dd_max)
    print()

    # 限定角色统计
    limited_rows = [row for row in h if row["类型"] == "限定"]
    total_limited = len(limited_rows)
    avg_lim_spent = sum(r["消耗抽"] for r in limited_rows) / total_limited if total_limited > 0 else 0
    wells_hit = sum(1 for r in limited_rows if "井" in r["流水"]) if total_limited > 0 else 0

    print("  【限定池统计】")
    print("    总限定池数:   %d 个" % total_limited)
    print("    平均消耗:     %.1f 抽/个" % avg_lim_spent)
    print("    吃井次数:     %d 次（占 %.1f%%）" % (wells_hit, wells_hit / total_limited * 100))
    print()

    # 补强统计
    total_supp = sum(r["补强数"] for r in h)
    print("  【红卡/积分补强统计】")
    print("    总补强次数:   %d 次（月均 %.2f 次）" % (total_supp, total_supp / (total_years * 12)))
    print()

    print("  【与 v10 的差异对比】")
    v10_monthly = 127.6
    v11_monthly = free_annual / 12 + SUBS_MONTHLY_PULLS
    print("    v10 月均抽: 127.6 抽（限定90抽）")
    print("    v11 月均抽: %.1f 抽（限定80抽）→ 减少 %.1f 抽/月（%.0f 抽/年）" % (
        v11_monthly, v10_monthly - v11_monthly, (v10_monthly - v11_monthly) * 12))
    print()

    # 保存 Excel
    out_dir  = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "..", "data")
    out_file = os.path.join(out_dir, "物华弥新_1000年模拟_v12_B档.xlsx")
    max_dd_val, final_pulls, final_reds = sim.save_excel(out_file)

    print()
    print("1000年后结余: %.0f 抽 | %.1f 红卡" % (final_pulls, final_reds))
    print("全局最大回撤: %d 抽（含库存积累效应，参考意义有限）" % max_dd_val)
    print("年内最大回撤P95: %d 抽（建议安全水位）" % dd_p95)

